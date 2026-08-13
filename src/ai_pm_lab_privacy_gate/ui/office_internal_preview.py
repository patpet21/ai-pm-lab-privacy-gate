from __future__ import annotations

import base64
import html
import re
from pathlib import Path

from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import column_index_from_string, get_column_letter
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QHeaderView,
    QLabel,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextBrowser,
    QVBoxLayout,
    QWidget,
)


TOKEN_PATTERN = re.compile(r"\[\[PG_([A-Z0-9_]+)_\d+\]\]")


class OfficeInternalPreview(QWidget):
    """Native, dependency-free preview for editable DOCX and XLSX files."""

    def __init__(self, colors: dict[str, str], parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.colors = colors
        self.tabs = QTabWidget()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self.tabs)

    def clear(self) -> None:
        self.tabs.clear()

    def focus_location(self, location: str) -> bool:
        """Open and select an Excel location such as ``Source Register!I4``."""
        match = re.fullmatch(r"(.+)!\$?([A-Z]+)\$?(\d+)", location.strip())
        if match is None:
            return False
        sheet_name, column_letters, row_text = match.groups()
        for index in range(self.tabs.count()):
            if self.tabs.tabText(index) != sheet_name:
                continue
            table = self.tabs.widget(index)
            if not isinstance(table, QTableWidget):
                return False
            row = int(row_text) - 1
            column = column_index_from_string(column_letters) - 1
            if row < 0 or column < 0 or row >= table.rowCount() or column >= table.columnCount():
                return False
            self.tabs.setCurrentIndex(index)
            table.setCurrentCell(row, column)
            item = table.item(row, column)
            if item is not None:
                table.scrollToItem(item, QAbstractItemView.ScrollHint.PositionAtCenter)
            return True
        return False

    def synchronize_with(self, other: "OfficeInternalPreview") -> None:
        """Keep matching Word/Excel preview panes aligned while scrolling."""
        for index in range(min(self.tabs.count(), other.tabs.count())):
            source = self.tabs.widget(index)
            target = other.tabs.widget(index)
            for accessor in ("horizontalScrollBar", "verticalScrollBar"):
                source_bar = getattr(source, accessor, lambda: None)()
                target_bar = getattr(target, accessor, lambda: None)()
                if source_bar is None or target_bar is None:
                    continue
                source_bar.valueChanged.connect(target_bar.setValue)
                target_bar.valueChanged.connect(source_bar.setValue)

    def load(self, path: str | Path, protected: bool = False) -> None:
        source = Path(path)
        self.clear()
        if source.suffix.lower() == ".docx":
            self._load_docx(source, protected)
        elif source.suffix.lower() == ".xlsx":
            self._load_xlsx(source, protected)
        else:
            raise ValueError("Internal Office preview supports DOCX and XLSX files.")

    def _load_docx(self, source: Path, protected: bool) -> None:
        document = Document(str(source))
        browser = QTextBrowser()
        browser.setOpenExternalLinks(False)
        browser.setObjectName("OfficeDocumentView")
        parts: list[str] = [
            "<html><body style='font-family:Segoe UI; color:#102A43; background:#fff; margin:20px'>"
        ]
        for block in document.iter_inner_content():
            if isinstance(block, Paragraph):
                parts.append(self._paragraph_html(block, protected))
            elif isinstance(block, Table):
                parts.append(self._table_html(block, protected))
        parts.append("</body></html>")
        browser.setHtml("".join(parts))
        self.tabs.addTab(browser, "Document")

    def _paragraph_html(self, paragraph: Paragraph, protected: bool) -> str:
        style_name = (paragraph.style.name if paragraph.style is not None else "").lower()
        tag = "p"
        if "title" in style_name:
            tag = "h1"
        elif "heading 1" in style_name:
            tag = "h2"
        elif "heading 2" in style_name:
            tag = "h3"
        elif "heading" in style_name:
            tag = "h4"
        fragments: list[str] = []
        for run in paragraph.runs:
            text = self._highlight(html.escape(run.text).replace("\n", "<br>"), protected)
            if run.bold:
                text = f"<b>{text}</b>"
            if run.italic:
                text = f"<i>{text}</i>"
            if run.underline:
                text = f"<u>{text}</u>"
            fragments.append(text)
            for blip in run._element.xpath(".//a:blip"):
                relation_id = blip.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
                if relation_id and relation_id in paragraph.part.related_parts:
                    part = paragraph.part.related_parts[relation_id]
                    mime = getattr(part, "content_type", "image/png")
                    encoded = base64.b64encode(part.blob).decode("ascii")
                    fragments.append(
                        f"<br><img src='data:{mime};base64,{encoded}' style='max-width:95%; max-height:360px'>"
                    )
        body = "".join(fragments) or "&nbsp;"
        return f"<{tag} style='margin:6px 0'>{body}</{tag}>"

    def _table_html(self, table: Table, protected: bool) -> str:
        rows = ["<table cellspacing='0' cellpadding='6' style='border-collapse:collapse; width:100%; margin:10px 0'>"]
        seen: set[object] = set()
        for row in table.rows:
            rows.append("<tr>")
            for cell in row.cells:
                marker = cell._tc
                if marker in seen:
                    continue
                seen.add(marker)
                content = "<br>".join(
                    self._highlight(html.escape(paragraph.text), protected)
                    for paragraph in cell.paragraphs
                ) or "&nbsp;"
                rows.append(
                    "<td style='border:1px solid #9FB3C8; vertical-align:top; background:#fff'>"
                    f"{content}</td>"
                )
            rows.append("</tr>")
        rows.append("</table>")
        return "".join(rows)

    def _highlight(self, escaped_text: str, protected: bool) -> str:
        if not protected:
            return escaped_text

        def replacement(match: re.Match[str]) -> str:
            entity = match.group(1)
            color = self.colors.get(entity, "#E7E9ED")
            return (
                f"<span style='background:{color}; color:#102A43; font-weight:600; "
                f"border-radius:3px; padding:1px 3px'>{match.group(0)}</span>"
            )

        return TOKEN_PATTERN.sub(replacement, escaped_text)

    def _load_xlsx(self, source: Path, protected: bool) -> None:
        workbook = load_workbook(source, data_only=False, keep_links=False)
        try:
            for worksheet in workbook.worksheets:
                table = self._worksheet_table(worksheet, protected)
                self.tabs.addTab(table, worksheet.title)
        finally:
            workbook.close()

    def _worksheet_table(self, worksheet, protected: bool) -> QWidget:
        max_row, max_column = self._used_bounds(worksheet)
        if max_row == 0 or max_column == 0:
            empty = QLabel("This worksheet is empty.")
            empty.setAlignment(Qt.AlignmentFlag.AlignCenter)
            return empty

        table = QTableWidget(max_row, max_column)
        table.setObjectName("OfficeSheetView")
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        table.setAlternatingRowColors(False)
        # A merge that crosses row 1 returns MergedCell placeholders for every
        # non-anchor column. Those placeholders intentionally have no
        # ``column_letter`` attribute, so derive labels from numeric indices.
        table.setHorizontalHeaderLabels(
            [get_column_letter(column) for column in range(1, max_column + 1)]
        )
        table.verticalHeader().setDefaultSectionSize(24)
        table.horizontalHeader().setDefaultSectionSize(110)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        for column in range(1, max_column + 1):
            dimension = worksheet.column_dimensions[get_column_letter(column)]
            if dimension.width:
                table.setColumnWidth(column - 1, max(55, min(360, int(dimension.width * 8))))
        for row in range(1, max_row + 1):
            height = worksheet.row_dimensions[row].height
            if height:
                table.setRowHeight(row - 1, max(20, min(180, int(height * 1.35))))
            for column in range(1, max_column + 1):
                cell = worksheet.cell(row, column)
                value = "" if cell.value is None else str(cell.value)
                item = QTableWidgetItem(value)
                item.setToolTip(value)
                if cell.font.bold:
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                foreground = self._openpyxl_color(cell.font.color)
                background = self._openpyxl_color(cell.fill.fgColor)
                if foreground:
                    item.setForeground(QColor(foreground))
                if background and cell.fill.fill_type:
                    item.setBackground(QColor(background))
                if protected:
                    token = TOKEN_PATTERN.search(value)
                    if token:
                        item.setBackground(QColor(self.colors.get(token.group(1), "#E7E9ED")))
                table.setItem(row - 1, column - 1, item)

        for merged in worksheet.merged_cells.ranges:
            if merged.max_row <= max_row and merged.max_col <= max_column:
                table.setSpan(
                    merged.min_row - 1,
                    merged.min_col - 1,
                    merged.max_row - merged.min_row + 1,
                    merged.max_col - merged.min_col + 1,
                )
        return table

    @staticmethod
    def _used_bounds(worksheet) -> tuple[int, int]:
        rows: list[int] = []
        columns: list[int] = []
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None or cell.has_style or cell.comment is not None:
                    rows.append(cell.row)
                    columns.append(cell.column)
        # Merged ranges can extend past the last independently populated cell.
        # Include their full footprint so the preview can preserve the span.
        for merged in worksheet.merged_cells.ranges:
            rows.append(merged.max_row)
            columns.append(merged.max_col)
        return (max(rows, default=0), max(columns, default=0))

    @staticmethod
    def _openpyxl_color(color) -> str | None:
        if color is None:
            return None
        if color.type == "rgb" and color.rgb:
            value = color.rgb[-6:]
            return f"#{value}" if value != "000000" or color.rgb.endswith("000000") else None
        if color.type == "indexed" and color.indexed is not None:
            index = int(color.indexed)
            if 0 <= index < len(COLOR_INDEX):
                return f"#{COLOR_INDEX[index][-6:]}"
        return None
