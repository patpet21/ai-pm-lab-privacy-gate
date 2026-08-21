from __future__ import annotations

import os
import re
from dataclasses import dataclass
from pathlib import Path

import pdfplumber
from PIL import ImageDraw
from docx import Document
from openpyxl import load_workbook
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

from ai_pm_lab_privacy_gate.domain.models import ReplacementMapping
from ai_pm_lab_privacy_gate.infrastructure.documents.office_service import OfficeDocumentService
from ai_pm_lab_privacy_gate.infrastructure.documents.pdf_service import PdfDocumentService


TOKEN_PATTERN = re.compile(r"\[\[PG_[A-Z0-9_]+_\d+\]\]")


@dataclass(frozen=True, slots=True)
class RestoreReport:
    output_path: Path
    restored_occurrences: int
    restored_tokens: tuple[str, ...]
    unknown_tokens: tuple[str, ...]


class DocumentRestoreService:
    """Restore Privacy Gate placeholders inside a protected result locally.

    The supplied document is treated as the layout source. The original source
    document is neither required nor loaded. Only the encrypted mapping selected
    from the local Library is used to replace matching tokens.
    """

    SUPPORTED_SUFFIXES = {".txt", ".pdf", ".docx", ".xlsx"}

    def restore(
        self,
        source: str | Path,
        mappings: tuple[ReplacementMapping, ...],
        destination: str | Path,
    ) -> RestoreReport:
        source_path = Path(source)
        if not source_path.is_file():
            raise FileNotFoundError(source_path)
        suffix = source_path.suffix.lower()
        if suffix not in self.SUPPORTED_SUFFIXES:
            raise ValueError("Restore supports TXT, PDF, DOCX and XLSX files.")
        output = Path(destination)
        if output.suffix.lower() != suffix:
            output = output.with_suffix(suffix)
        output.parent.mkdir(parents=True, exist_ok=True)
        replacements = {item.token: item.original_text for item in mappings}
        if suffix == ".txt":
            text = source_path.read_text(encoding="utf-8-sig")
            restored, count, present, unknown = self._restore_text(text, replacements)
            output.write_text(restored, encoding="utf-8")
        elif suffix == ".docx":
            count, present, unknown = self._restore_docx(source_path, output, replacements)
        elif suffix == ".xlsx":
            count, present, unknown = self._restore_xlsx(source_path, output, replacements)
        else:
            count, present, unknown = self._restore_pdf(source_path, output, replacements)
        return RestoreReport(
            output_path=output,
            restored_occurrences=count,
            restored_tokens=tuple(sorted(present)),
            unknown_tokens=tuple(sorted(unknown)),
        )

    @staticmethod
    def extract_text(source: str | Path) -> str:
        path = Path(source)
        suffix = path.suffix.lower()
        if suffix == ".txt":
            return path.read_text(encoding="utf-8-sig")
        if suffix == ".pdf":
            with pdfplumber.open(path) as pdf:
                return "\n\n".join(page.extract_text() or "" for page in pdf.pages)
        if suffix == ".docx":
            document = Document(str(path))
            return "\n".join(
                part.paragraph.text
                for part in OfficeDocumentService._iter_docx_parts(document)
                if part.paragraph.text
            )
        if suffix == ".xlsx":
            workbook = load_workbook(path, data_only=False, keep_links=False)
            try:
                return "\n".join(
                    OfficeDocumentService._excel_part_text(part)
                    for part in OfficeDocumentService._iter_excel_parts(workbook)
                )
            finally:
                workbook.close()
        raise ValueError("Restore supports TXT, PDF, DOCX and XLSX files.")

    @staticmethod
    def _restore_text(
        text: str, replacements: dict[str, str]
    ) -> tuple[str, int, set[str], set[str]]:
        present = set(TOKEN_PATTERN.findall(text))
        unknown = present.difference(replacements)
        count = 0
        for token, original in replacements.items():
            occurrences = text.count(token)
            if occurrences:
                text = text.replace(token, original)
                count += occurrences
        return text, count, present.intersection(replacements), unknown

    def _restore_docx(
        self, source: Path, output: Path, replacements: dict[str, str]
    ) -> tuple[int, set[str], set[str]]:
        document = Document(str(source))
        count = 0
        present: set[str] = set()
        for part in OfficeDocumentService._iter_docx_parts(document):
            paragraph = part.paragraph
            text = paragraph.text
            matches = list(TOKEN_PATTERN.finditer(text))
            present.update(match.group(0) for match in matches)
            for match in reversed(matches):
                replacement = replacements.get(match.group(0))
                if replacement is None:
                    continue
                OfficeDocumentService._replace_paragraph_range(
                    paragraph, match.start(), match.end(), replacement
                )
                count += 1
        OfficeDocumentService._clear_docx_identity(document)
        document.save(str(output))
        return count, present.intersection(replacements), present.difference(replacements)

    def _restore_xlsx(
        self, source: Path, output: Path, replacements: dict[str, str]
    ) -> tuple[int, set[str], set[str]]:
        workbook = load_workbook(source, data_only=False, keep_links=False)
        count = 0
        present: set[str] = set()
        try:
            for part in OfficeDocumentService._iter_excel_parts(workbook):
                text = OfficeDocumentService._excel_part_text(part)
                restored, item_count, item_present, _unknown = self._restore_text(text, replacements)
                present.update(TOKEN_PATTERN.findall(text))
                if not item_count:
                    continue
                count += item_count
                if part.component == "comment" and part.cell.comment is not None:
                    part.cell.comment.text = restored
                else:
                    part.cell.value = restored
            OfficeDocumentService._clear_xlsx_identity(workbook)
            workbook.save(output)
        finally:
            workbook.close()
        return count, present.intersection(replacements), present.difference(replacements)

    def _restore_pdf(
        self, source: Path, output: Path, replacements: dict[str, str]
    ) -> tuple[int, set[str], set[str]]:
        count = 0
        present: set[str] = set()
        unresolved: list[str] = []
        temporary = output.with_name(f".{output.stem}.restoring-{os.getpid()}.pdf")
        temporary.unlink(missing_ok=True)
        with pdfplumber.open(source) as pdf:
            writer = canvas.Canvas(str(temporary))
            try:
                for page in pdf.pages:
                    page_text = page.extract_text() or ""
                    page_tokens = set(TOKEN_PATTERN.findall(page_text))
                    present.update(page_tokens)
                    rendered = page.to_image(resolution=144, antialias=True).original.convert("RGB")
                    draw = ImageDraw.Draw(rendered)
                    scale_x = rendered.width / float(page.width)
                    scale_y = rendered.height / float(page.height)
                    for token in sorted(page_tokens, key=len, reverse=True):
                        original = replacements.get(token)
                        if original is None:
                            continue
                        matches = page.search(re.escape(token), regex=True, case=True)
                        if not matches:
                            unresolved.append(token)
                            continue
                        for match in matches:
                            box = (
                                max(0, int(match["x0"] * scale_x) - 5),
                                max(0, int(match["top"] * scale_y) - 5),
                                min(rendered.width, int(match["x1"] * scale_x) + 5),
                                min(rendered.height, int(match["bottom"] * scale_y) + 5),
                            )
                            # Cover the protected token chip completely, then
                            # draw the original value without a visible border.
                            # On ordinary white document backgrounds this reads
                            # like normal document text instead of a form field.
                            draw.rectangle(box, fill="#FFFFFF")
                            PdfDocumentService._draw_fitted_text(
                                draw,
                                box,
                                original.replace("\n", " "),
                                align="left",
                            )
                            count += 1
                    writer.setPageSize((float(page.width), float(page.height)))
                    writer.drawImage(
                        ImageReader(rendered), 0, 0,
                        width=float(page.width), height=float(page.height),
                    )
                    writer.showPage()
                writer.save()
            except Exception:
                writer._doc = None
                raise
        if unresolved:
            temporary.unlink(missing_ok=True)
            raise ValueError(
                "Unable to safely locate these placeholders in the PDF: "
                + ", ".join(sorted(set(unresolved))[:8])
            )
        os.replace(temporary, output)
        return count, present.intersection(replacements), present.difference(replacements)
